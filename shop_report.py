import argparse
import hashlib
import hmac
import json
import logging
import re
import sys
import time
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import requests

BASE_DIR = Path(__file__).resolve().parent
KEYS_FILE = BASE_DIR / "tk_shop_keys.txt"
CONFIG_FILE = BASE_DIR / "config.json"
API_BASE = "https://open-api.tiktokglobalshop.com"
AUTH_BASE = "https://auth.tiktok-shops.com"
RECORDINGS_DIR = Path("F:/tiktok-recordings")
EXCEL_PATH = RECORDINGS_DIR / "店铺日报.xlsx"
# Mexico City is UTC-6 (no DST since 2022); order-day boundaries follow shop local time.
MX_TZ = timezone(timedelta(hours=-6))

ORDER_STATUS_ZH = {
    "UNPAID": "待付款",
    "ON_HOLD": "留置中",
    "AWAITING_SHIPMENT": "待发货",
    "PARTIALLY_SHIPPING": "部分发货",
    "AWAITING_COLLECTION": "待揽收",
    "IN_TRANSIT": "运输中",
    "DELIVERED": "已送达",
    "COMPLETED": "已完成",
    "CANCELLED": "已取消",
}


def load_keys() -> Dict[str, str]:
    if not KEYS_FILE.exists():
        logging.error("Keys file missing: %s", KEYS_FILE)
        sys.exit(1)
    return {
        k.strip(): v.strip()
        for k, v in (
            line.split("=", 1)
            for line in KEYS_FILE.read_text(encoding="utf-8-sig").strip().splitlines()
            if "=" in line
        )
    }


def save_keys(keys: Dict[str, str]) -> None:
    KEYS_FILE.write_text("\n".join(f"{k}={v}" for k, v in keys.items()) + "\n", encoding="utf-8")


def refresh_token_if_needed(keys: Dict[str, str]) -> Dict[str, str]:
    expire_at = int(keys.get("access_token_expire_in", "0") or 0)
    if expire_at - time.time() > 24 * 3600:
        return keys
    logging.info("Access token expiring soon; refreshing")
    resp = requests.get(
        f"{AUTH_BASE}/api/v2/token/refresh",
        params={
            "app_key": keys["app_key"],
            "app_secret": keys["app_secret"],
            "refresh_token": keys["refresh_token"],
            "grant_type": "refresh_token",
        },
        timeout=30,
    ).json()
    data = resp.get("data") or {}
    if not data.get("access_token"):
        logging.error("Token refresh failed: %s", resp)
        sys.exit(1)
    keys["access_token"] = data["access_token"]
    keys["refresh_token"] = data["refresh_token"]
    keys["access_token_expire_in"] = str(data.get("access_token_expire_in"))
    keys["refresh_token_expire_in"] = str(data.get("refresh_token_expire_in"))
    save_keys(keys)
    logging.info("Token refreshed")
    return keys


def api_call(
    keys: Dict[str, str],
    method: str,
    path: str,
    params: Optional[Dict[str, str]] = None,
    body: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    secret = keys["app_secret"]
    params = dict(params or {})
    params["app_key"] = keys["app_key"]
    params["timestamp"] = str(int(time.time()))
    body_str = json.dumps(body) if body is not None else ""
    base = secret + path + "".join(k + params[k] for k in sorted(params)) + body_str + secret
    params["sign"] = hmac.new(secret.encode(), base.encode(), hashlib.sha256).hexdigest()
    for attempt in range(3):
        resp = requests.request(
            method,
            API_BASE + path,
            params=params,
            headers={"x-tts-access-token": keys["access_token"], "content-type": "application/json"},
            data=body_str if body is not None else None,
            timeout=60,
        ).json()
        if resp.get("code") == 0:
            return resp.get("data") or {}
        if attempt < 2:
            import time as _t; _t.sleep(3)
    raise RuntimeError(f"API {path} failed: {resp.get('code')} {resp.get('message')}")
    return resp.get("data") or {}


def get_shop_cipher(keys: Dict[str, str]) -> str:
    data = api_call(keys, "GET", "/authorization/202309/shops")
    shops = data.get("shops") or []
    if not shops:
        raise RuntimeError("No authorized shops returned")
    return shops[0]["cipher"]


def fetch_orders(keys: Dict[str, str], cipher: str, start_ts: int, end_ts: int) -> List[Dict[str, Any]]:
    orders: List[Dict[str, Any]] = []
    page_token = ""
    while True:
        params = {"shop_cipher": cipher, "page_size": "50"}
        if page_token:
            params["page_token"] = page_token
        data = api_call(
            keys,
            "POST",
            "/order/202309/orders/search",
            params,
            {"create_time_ge": start_ts, "create_time_lt": end_ts},
        )
        orders.extend(data.get("orders") or [])
        page_token = data.get("next_page_token") or ""
        if not page_token:
            break
        import time; time.sleep(1)
    return orders


def collect_live_stats(day: str) -> List[Dict[str, Any]]:
    # stats.json started_at is recorded in Beijing time (UTC+8); convert to
    # Mexico time (UTC-6) so live sessions align with the report's order day.
    stats = []
    if not RECORDINGS_DIR.exists():
        return stats
    for path in RECORDINGS_DIR.rglob("*.stats.json"):
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            started_bj = datetime.strptime(data.get("started_at", ""), "%Y-%m-%d %H:%M:%S")
            started_mx = started_bj.replace(tzinfo=timezone(timedelta(hours=8))).astimezone(MX_TZ)
            if started_mx.strftime("%Y-%m-%d") == day:
                data["started_at_mx"] = started_mx.strftime("%H:%M")
                stats.append(data)
        except Exception:
            logging.warning("Bad stats file: %s", path)
    stats.sort(key=lambda s: s.get("started_at", ""))
    return stats


MX_STATE_RANGES = [
    (0, 16, "墨西哥城"), (20, 20, "阿瓜斯卡连特斯"), (21, 22, "下加州"), (23, 23, "南下加州"),
    (24, 24, "坎佩切"), (25, 27, "科阿韦拉"), (28, 28, "科利马"), (29, 30, "恰帕斯"),
    (31, 33, "奇瓦瓦"), (34, 35, "杜兰戈"), (36, 38, "瓜纳华托"), (39, 41, "格雷罗"),
    (42, 43, "伊达尔戈"), (44, 49, "哈利斯科"), (50, 57, "墨西哥州"), (58, 61, "米却肯"),
    (62, 62, "莫雷洛斯"), (63, 63, "纳亚里特"), (64, 67, "新莱昂"), (68, 71, "瓦哈卡"),
    (72, 75, "普埃布拉"), (76, 76, "克雷塔罗"), (77, 77, "金塔纳罗奥"), (78, 79, "圣路易斯波托西"),
    (80, 82, "锡那罗亚"), (83, 85, "索诺拉"), (86, 86, "塔巴斯科"), (87, 89, "塔毛利帕斯"),
    (90, 90, "特拉斯卡拉"), (91, 96, "韦拉克鲁斯"), (97, 97, "尤卡坦"), (98, 99, "萨卡特卡斯"),
]


def state_from_zip(postal_code: str) -> Optional[str]:
    digits = (postal_code or "").strip()[:2]
    if not digits.isdigit():
        return None
    prefix = int(digits)
    for low, high, name in MX_STATE_RANGES:
        if low <= prefix <= high:
            return name
    return None


def hourly_sparkline(hour_count: Dict[int, int]) -> str:
    blocks = "▁▂▃▄▅▆▇█"
    high = max(hour_count.values()) if hour_count else 0
    if high <= 0:
        return ""
    return "".join(
        blocks[0] if hour_count.get(h, 0) == 0 else blocks[max(1, int(hour_count[h] / high * (len(blocks) - 1)))]
        for h in range(24)
    )


def build_report(
    day: str,
    orders: List[Dict[str, Any]],
    live_stats: List[Dict[str, Any]],
    prev_day_orders: Optional[int] = None,
) -> str:
    valid = [o for o in orders if o.get("status") != "CANCELLED"]
    cancelled = [o for o in orders if o.get("status") == "CANCELLED"]

    product_qty: Dict[str, int] = defaultdict(int)
    size_qty: Dict[str, int] = defaultdict(int)
    color_qty: Dict[str, int] = defaultdict(int)
    item_total = 0
    for order in valid:
        for item in order.get("line_items") or []:
            name = item.get("product_name") or "未知商品"
            product_qty[name] += 1
            item_total += 1
            sku = (item.get("sku_name") or "").strip()
            if "," in sku:
                color, size = (part.strip() for part in sku.rsplit(",", 1))
                if color:
                    color = re.sub(r"\s*\([^)]*\)", "", color).strip() or color
                    color_qty[color] += 1
                if size:
                    size_qty[size.upper()] += 1
            elif sku:
                size_qty[sku.upper()] += 1
    top_products = sorted(product_qty.items(), key=lambda kv: -kv[1])[:5]

    status_count: Dict[str, int] = defaultdict(int)
    for order in orders:
        status_count[order.get("status", "UNKNOWN")] += 1

    hour_count: Dict[int, int] = defaultdict(int)
    for order in valid:
        try:
            hour_count[datetime.fromtimestamp(int(order.get("create_time") or 0), MX_TZ).hour] += 1
        except (TypeError, ValueError):
            pass

    cod_count = sum(1 for o in valid if o.get("is_cod"))
    payment_count: Dict[str, int] = defaultdict(int)
    provider_count: Dict[str, int] = defaultdict(int)
    state_count: Dict[str, int] = defaultdict(int)
    sample_count = 0
    for order in valid:
        method = order.get("payment_method_name") or "未知"
        payment_count[method] += 1
        provider = order.get("shipping_provider") or "未分配"
        provider_count[provider] += 1
        state = state_from_zip((order.get("recipient_address") or {}).get("postal_code") or "")
        if state:
            state_count[state] += 1
        if order.get("is_sample_order"):
            sample_count += 1

    cancel_reasons: Dict[str, int] = defaultdict(int)
    for order in cancelled:
        reason = (order.get("cancel_reason") or "未注明").strip()
        cancel_reasons[reason if len(reason) <= 30 else reason[:30] + "…"] += 1

    lines = [f"📊 店铺日报 {day}（墨西哥时间）", ""]
    trend = ""
    if prev_day_orders is not None and prev_day_orders > 0:
        delta = (len(orders) - prev_day_orders) / prev_day_orders * 100
        arrow = "↑" if delta >= 0 else "↓"
        trend = f"（环比{arrow}{abs(delta):.0f}%）"
    lines.append(f"订单: {len(orders)} 单{trend}｜有效 {len(valid)} / 取消 {len(cancelled)}｜商品 {item_total} 件")
    if sample_count:
        lines.append(f"其中样品单: {sample_count} 单")
    if status_count:
        status_text = " | ".join(
            f"{ORDER_STATUS_ZH.get(k, k)} {v}" for k, v in sorted(status_count.items(), key=lambda kv: -kv[1])
        )
        lines.append(f"状态: {status_text}")

    if size_qty or color_qty:
        lines.append("")
        lines.append("— 商品维度 —")
    if size_qty:
        size_order = {"XS": 0, "S": 1, "M": 2, "L": 3, "XL": 4, "2XL": 5, "3XL": 6, "4XL": 7, "5XL": 8}
        ordered = sorted(size_qty.items(), key=lambda kv: size_order.get(kv[0], 99))
        lines.append("尺码: " + " | ".join(f"{s}×{q}" for s, q in ordered))
    if color_qty:
        ordered_colors = sorted(color_qty.items(), key=lambda kv: -kv[1])
        lines.append("颜色: " + " | ".join(f"{c}×{q}" for c, q in ordered_colors[:6]))
    if top_products:
        lines.append("热卖:")
        for name, qty in top_products:
            short = name if len(name) <= 38 else name[:38] + "…"
            lines.append(f"  · {short} × {qty} 件")

    lines.append("")
    lines.append("— 买家维度 —")
    if valid:
        lines.append(f"货到付款: {cod_count} 单（{cod_count / len(valid) * 100:.0f}%）")
    if payment_count:
        lines.append("支付方式: " + " | ".join(f"{m}×{q}" for m, q in sorted(payment_count.items(), key=lambda kv: -kv[1])[:4]))
    if state_count:
        top_states = sorted(state_count.items(), key=lambda kv: -kv[1])[:6]
        lines.append("地区: " + " | ".join(f"{s}×{q}" for s, q in top_states))
    if hour_count:
        peak_hour = max(hour_count, key=hour_count.get)
        lines.append(f"时段: {hourly_sparkline(hour_count)} (0-23时)")
        lines.append(f"下单高峰: {peak_hour:02d}:00-{peak_hour + 1:02d}:00（{hour_count[peak_hour]} 单）")

    if provider_count or cancel_reasons:
        lines.append("")
        lines.append("— 履约维度 —")
    if provider_count:
        lines.append("物流商: " + " | ".join(f"{p}×{q}" for p, q in sorted(provider_count.items(), key=lambda kv: -kv[1])[:4]))
    if cancel_reasons:
        lines.append("取消原因:")
        for reason, qty in sorted(cancel_reasons.items(), key=lambda kv: -kv[1]):
            lines.append(f"  · {reason} × {qty}")

    lines.append("")
    lines.append("— 直播维度 —")
    if live_stats:
        total_minutes = sum(int(s.get("duration_minutes") or 0) for s in live_stats)
        streamers = len({s.get("username") for s in live_stats})
        lines.append(f"直播: {len(live_stats)} 场｜{streamers} 位达人｜累计 {total_minutes} 分钟")
        for s in live_stats:
            lines.append(
                f"  · @{s.get('username')} 墨西哥{s.get('started_at_mx', '?')}开播 "
                f"{s.get('duration_minutes')}分钟 均在线{s.get('avg_viewers')}人 峰值{s.get('peak_viewers')}人"
            )
    else:
        lines.append("直播: 0 场（无录制记录）")
    return "\n".join(lines)


def write_excel(day: str, orders: List[Dict[str, Any]]) -> Optional[Path]:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        logging.warning("openpyxl not installed; skipping Excel export")
        return None

    headers = ["日期", "订单号", "状态", "金额", "币种", "商品", "下单时间(墨西哥)"]
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        for row in list(ws.iter_rows(min_row=2)):
            if row[0].value == day:
                ws.delete_rows(row[0].row, 1)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "订单明细"
        ws.append(headers)

    for order in orders:
        items = "; ".join((i.get("product_name") or "?") for i in (order.get("line_items") or []))
        created = datetime.fromtimestamp(int(order.get("create_time") or 0), MX_TZ).strftime("%Y-%m-%d %H:%M")
        ws.append(
            [
                day,
                str(order.get("id")),
                ORDER_STATUS_ZH.get(order.get("status", ""), order.get("status")),
                float(order.get("payment", {}).get("total_amount") or 0),
                order.get("payment", {}).get("currency") or "MXN",
                items[:200],
                created,
            ]
        )
    wb.save(EXCEL_PATH)
    return EXCEL_PATH


def send_feishu(text: str) -> None:
    try:
        webhook = json.loads(CONFIG_FILE.read_text(encoding="utf-8-sig")).get("feishu_webhook")
    except Exception:
        webhook = None
    if not webhook:
        logging.warning("No feishu_webhook in config.json; printing report only")
        return
    resp = requests.post(webhook, json={"msg_type": "text", "content": {"text": text}}, timeout=30)
    if resp.status_code != 200 or resp.json().get("code") not in (0, None):
        logging.error("Feishu send failed: %s %s", resp.status_code, resp.text[:200])
    else:
        logging.info("Report sent to Feishu")


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s", datefmt="%H:%M:%S")
    parser = argparse.ArgumentParser(description="TikTok Shop daily report")
    parser.add_argument("--date", help="Report date YYYY-MM-DD in Mexico time (default: yesterday)")
    parser.add_argument("--no-send", action="store_true", help="Print report without sending to Feishu")
    args = parser.parse_args()

    if args.date:
        day_start = datetime.strptime(args.date, "%Y-%m-%d").replace(tzinfo=MX_TZ)
    else:
        now_mx = datetime.now(MX_TZ)
        day_start = (now_mx - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    day = day_start.strftime("%Y-%m-%d")
    day_end = day_start + timedelta(days=1)

    keys = refresh_token_if_needed(load_keys())
    cipher = get_shop_cipher(keys)
    logging.info("Fetching orders for %s (MX time)", day)
    orders = fetch_orders(keys, cipher, int(day_start.timestamp()), int(day_end.timestamp()))
    logging.info("Fetched %d order(s)", len(orders))
    prev_start = day_start - timedelta(days=1)
    try:
        prev_day_orders = len(fetch_orders(keys, cipher, int(prev_start.timestamp()), int(day_start.timestamp())))
    except Exception:
        prev_day_orders = None
    live_stats = collect_live_stats(day)

    report = build_report(day, orders, live_stats, prev_day_orders)
    print(report)

    excel = write_excel(day, orders)
    if excel:
        logging.info("Excel updated: %s", excel)
    if not args.no_send:
        send_feishu(report)
    return 0


if __name__ == "__main__":
    sys.exit(main())
